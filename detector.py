# detector.py
"""
Detecção de extensão por assinatura ("magic number") + heurísticas.
- Suporte a dezenas de formatos (imagens, áudio/vídeo, compactados, docs, fontes etc.)
- Inspeção de ZIP para diferenciar docx/xlsx/pptx/apk/jar/epub/odt/ods/odp.
- Gera .zip com arquivos processados e:
    - XLSX formatado (cabeçalho preto, linhas cinza)
    - CSV (UTF-8 BOM + ; para Excel pt-BR) OPCIONAL
- Fornece resumo por formato com descrição e software recomendado.
"""

from __future__ import annotations
from pathlib import Path
from typing import Dict, List, Tuple
import io
import os
import zipfile
import tempfile
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ------------------------------------------------------------------
# Descrições e recomendações (usadas no resumo)
# ------------------------------------------------------------------
FORMAT_INFO: Dict[str, Tuple[str, str]] = {
    ".jpg": ("Imagem JPEG (compactada, foto)", "IrfanView / XnView MP / Fotos (Win)"),
    ".png": ("Imagem PNG (sem perda, transparência)", "IrfanView / XnView MP"),
    ".gif": ("Imagem GIF (paleta, pode ser animada)", "XnView MP"),
    ".bmp": ("Imagem BMP (bitmap sem compressão)", "IrfanView / XnView MP"),
    ".tiff": ("Imagem TIFF (multi-página, sem/perda)", "IrfanView / XnView MP"),
    ".psd": ("Adobe Photoshop Document", "Photopea (web) / GIMP"),
    ".webp": ("Imagem WebP (compactada)", "IrfanView / navegadores"),
    ".heic": ("Imagem HEIC/HEIF (alta eficiência)", "CopyTrans HEIC (Win) / XnView MP"),
    ".ico": ("Ícone (favicon/atalho)", "IrfanView / XnView MP"),

    ".pdf": ("Documento PDF", "Adobe Reader / SumatraPDF"),
    ".docx": ("Documento Word (Office Open XML)", "LibreOffice Writer"),
    ".xlsx": ("Planilha Excel (Office Open XML)", "LibreOffice Calc"),
    ".pptx": ("Apresentação PowerPoint (Office Open XML)", "LibreOffice Impress"),
    ".odt": ("Documento ODT (OpenDocument)", "LibreOffice Writer"),
    ".ods": ("Planilha ODS (OpenDocument)", "LibreOffice Calc"),
    ".odp": ("Apresentação ODP (OpenDocument)", "LibreOffice Impress"),
    ".epub": ("eBook EPUB", "Calibre / SumatraPDF"),

    ".mp3": ("Áudio MP3", "VLC / foobar2000"),
    ".flac": ("Áudio FLAC (sem perda)", "VLC / foobar2000"),
    ".ogg": ("Áudio OGG/Vorbis", "VLC"),
    ".opus": ("Áudio OPUS", "VLC / foobar2000"),
    ".wav": ("Áudio WAV (RIFF)", "VLC / Audacity"),

    ".mp4": ("Vídeo MP4 (H.264/H.265)", "VLC"),
    ".mkv": ("Vídeo Matroska MKV", "VLC"),
    ".webm": ("Vídeo WebM", "VLC"),
    ".avi": ("Vídeo AVI (RIFF)", "VLC"),

    ".zip": ("Arquivo ZIP (compactado)", "7-Zip / PeaZip"),
    ".7z": ("Arquivo 7-Zip", "7-Zip / PeaZip"),
    ".gz": ("GZip (compressão stream)", "7-Zip / PeaZip"),
    ".bz2": ("BZip2 (compressão stream)", "7-Zip / PeaZip"),
    ".xz": ("XZ (compressão stream)", "7-Zip / PeaZip"),
    ".rar": ("RAR", "PeaZip (gratuito)"),

    ".json": ("JSON (texto estruturado)", "VS Code / Notepad++"),
    ".xml": ("XML (texto estruturado)", "VS Code / Notepad++"),
    ".html": ("HTML (página web)", "Navegador / VS Code"),
    ".txt": ("Texto simples)", "Notepad++ / VS Code"),

    ".sqlite": ("Banco de dados SQLite)", "DB Browser for SQLite"),
    ".exe": ("Executável Windows (PE)", "—"),
    ".jar": ("Java JAR (ZIP com classes)", "Java Runtime / 7-Zip"),
    ".apk": ("Android APK (ZIP)", "Android Studio / APKTool"),
    ".woff": ("Fonte WOFF)", "Visores de fontes / Brotli"),
    ".woff2": ("Fonte WOFF2)", "Visores de fontes / Brotli"),
    ".ttf": ("Fonte TrueType)", "Visualizador de fontes do SO"),
    ".otf": ("Fonte OpenType)", "Visualizador de fontes do SO"),
}

# Se já tiver alguma dessas extensões, não renomeamos
EXTENSOES_CONHECIDAS = set(FORMAT_INFO.keys()) | {
    ".pages", ".numbers", ".key", ".dex", ".oat", ".art", ".aab", ".log",
    ".db", ".pb", ".proto", ".nomedia", ".gdoc", ".gsheet"
}

# ------------------------------------------------------------------
# Tabela de magic numbers (prefixos)
# ------------------------------------------------------------------
MAGIC_PREFIX: Dict[bytes, str] = {
    b"\xFF\xD8\xFF": ".jpg",
    b"\x89PNG\r\n\x1a\n": ".png",
    b"GIF87a": ".gif",
    b"GIF89a": ".gif",
    b"BM": ".bmp",
    b"\x49\x49\x2A\x00": ".tiff",
    b"\x4D\x4D\x00\x2A": ".tiff",
    b"8BPS": ".psd",

    b"fLaC": ".flac",
    b"ID3": ".mp3",
    b"OggS": ".ogg",                 # container (detalhamos abaixo)
    b"RIFF": "riff",                 # WAV/AVI/WebP (detalhamos)
    b"\x1A\x45\xDF\xA3": "ebml",     # MKV/WebM (detalhamos)

    b"%PDF": ".pdf",
    b"\x50\x4B\x03\x04": "zip",      # docx/xlsx/pptx/apk/jar/epub/odf…
    b"\x37\x7A\xBC\xAF\x27\x1C": ".7z",
    b"\x1F\x8B\x08": ".gz",
    b"BZh": ".bz2",
    b"\xFD7zXZ\x00": ".xz",
    b"Rar!\x1A\x07\x00": ".rar",     # RAR4
    b"Rar!\x1A\x07\x01\x00": ".rar", # RAR5

    b"\x7FELF": ".elf",
    b"MZ": ".exe",

    b"wOFF": ".woff",
    b"wOF2": ".woff2",
    b"\x00\x01\x00\x00": ".ttf",
    b"OTTO": ".otf",
}

# ---------------- Heurísticas auxiliares ----------------
def _is_text(data: bytes) -> bool:
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        return False
    return all((32 <= ord(c) <= 126) or c in "\r\n\t" for c in text)

def _riff_detail(data: bytes) -> str | None:
    if b"WAVE" in data[:64]:
        return ".wav"
    if b"AVI " in data[:64]:
        return ".avi"
    if b"WEBP" in data[:64] or b"WEBPVP8" in data[:256]:
        return ".webp"
    return None

def _ebml_detail(data: bytes) -> str | None:
    head = data[:4096].lower()
    if b"webm" in head:
        return ".webm"
    if b"matroska" in head:
        return ".mkv"
    return ".mkv"

def _mp4_family(data: bytes) -> str | None:
    if b"ftyp" in data[4:16]:
        return ".mp4"  # simplificado
    return None

def _heif_family(data: bytes) -> str | None:
    head = data[:32]
    if b"ftypheic" in head or b"ftypheif" in head or b"ftypmif1" in head or b"ftypmsf1" in head:
        return ".heic"
    return None

def _ogg_detail(data: bytes) -> str | None:
    if b"OpusHead" in data[:64]:
        return ".opus"
    return ".ogg"

def _zip_detail_all(data: bytes) -> str | None:
    """Diferencia docx/xlsx/pptx/apk/jar/epub/odt/ods/odp analisando o ZIP em memória."""
    try:
        bio = io.BytesIO(data)
        with zipfile.ZipFile(bio) as z:
            names = set(z.namelist())

            # Office Open XML
            if any(n.startswith("word/") for n in names):
                return ".docx"
            if any(n.startswith("xl/") for n in names):
                return ".xlsx"
            if any(n.startswith("ppt/") for n in names):
                return ".pptx"

            # APK
            if "AndroidManifest.xml" in names and "classes.dex" in names:
                return ".apk"

            # JAR
            if "META-INF/MANIFEST.MF" in names and any(n.endswith(".class") for n in names):
                return ".jar"

            # EPUB
            if "mimetype" in names:
                try:
                    with z.open("mimetype") as mf:
                        mt = mf.read(80)
                        if b"application/epub+zip" in mt:
                            return ".epub"
                except Exception:
                    pass

            # OpenDocument
            if "mimetype" in names:
                try:
                    with z.open("mimetype") as mf:
                        mt = mf.read(160)
                        if b"application/vnd.oasis.opendocument.text" in mt:
                            return ".odt"
                        if b"application/vnd.oasis.opendocument.spreadsheet" in mt:
                            return ".ods"
                        if b"application/vnd.oasis.opendocument.presentation" in mt:
                            return ".odp"
                except Exception:
                    pass

            return ".zip"
    except Exception:
        return ".zip"

def detectar_extensao(data: bytes) -> str | None:
    # famílias especiais primeiro
    ext = _mp4_family(data)
    if ext:
        return ext
    ext = _heif_family(data)
    if ext:
        return ext

    # prefixos diretos
    for magic, label in MAGIC_PREFIX.items():
        if data.startswith(magic):
            if label == "riff":
                return _riff_detail(data)
            if label == "ebml":
                return _ebml_detail(data)
            if label == ".ogg":
                return _ogg_detail(data)
            if label == "zip":
                return _zip_detail_all(data)
            return label

    # Heurísticas: JSON/HTML/XML/TXT
    head = data.lstrip()
    if head.startswith(b"{") or head.startswith(b"["):
        try:
            head.decode("utf-8")
            return ".json"
        except Exception:
            pass
    if head.lower().startswith(b"<!doctype html") or head.lower().startswith(b"<html"):
        return ".html"
    if head.startswith(b"<?xml"):
        return ".xml"
    if _is_text(data):
        return ".txt"

    return None

# ----------------------------------------------------------------------
# XLSX (tabela preta/cinza)
# ----------------------------------------------------------------------
def _escrever_excel_formatado(relatorio: List[Tuple[str, str, str]], xlsx_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    headers = ["arquivo_original", "saida", "extensão detectada"]

    fill_header = PatternFill("solid", fgColor="111111")
    font_header = Font(bold=True, color="FFFFFF")
    fill_row = PatternFill("solid", fgColor="E5E5E5")
    border_thin = Border(
        left=Side(style="thin", color="444444"),
        right=Side(style="thin", color="444444"),
        top=Side(style="thin", color="444444"),
        bottom=Side(style="thin", color="444444"),
    )
    align_left = Alignment(horizontal="left", vertical="center")

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_left
        cell.border = border_thin

    for (orig, saida, ext) in relatorio:
        ws.append([orig, saida, ext or ""])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.fill = fill_row
            cell.alignment = align_left
            cell.border = border_thin

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for c in col_cells:
            val = str(c.value) if c.value is not None else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(xlsx_path)

def _zip_dir(folder: Path, base_name: str) -> str:
    out = Path.cwd() / f"{base_name}.zip"
    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for p in folder.rglob("*"):
            if p.is_file():
                z.write(p, arcname=str(p.relative_to(folder)))
    return str(out)

# ----------------------------------------------------------------------
# Pipelines
# ----------------------------------------------------------------------
def processar_arquivos(
    paths: List[Path],
    gerar_csv: bool = False,
) -> Tuple[str, List[Tuple[str, str, str]], Dict[str, int]]:
    """
    Recebe uma lista de arquivos (não-zip), detecta extensões e gera:
    - pasta temporária com arquivos renomeados/copiados
    - .xlsx formatado e, opcionalmente, .csv
    - .zip final com tudo
    """
    temp_out = Path(tempfile.mkdtemp())
    relatorio: List[Tuple[str, str, str]] = []
    counts: Dict[str, int] = {}

    for p in paths:
        data = Path(p).read_bytes()
        head = data[:4096]
        nome = p.name
        _base, ext_atual = os.path.splitext(nome)

        if ext_atual and ext_atual.lower() in EXTENSOES_CONHECIDAS:
            (temp_out / nome).write_bytes(data)
            ext = ext_atual.lower()
        else:
            ext = detectar_extensao(head) or ""
            novo_nome = nome + ext if ext and not ext_atual else nome
            (temp_out / novo_nome).write_bytes(data)
            nome = novo_nome

        relatorio.append((p.name, nome, ext))
        if ext:
            counts[ext] = counts.get(ext, 0) + 1

    # XLSX SEMPRE
    xlsx_path = temp_out / "relatorio_renomeacao.xlsx"
    _escrever_excel_formatado(relatorio, xlsx_path)

    # CSV OPCIONAL
    if gerar_csv:
        csv_path = temp_out / "relatorio_renomeacao.csv"
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
            writer.writerow(["arquivo_original", "saida", "ext_detectada"])
            for a, b, c in relatorio:
                writer.writerow([a, b, c])

    zip_path = _zip_dir(temp_out, "renomeados_resultado")
    return zip_path, relatorio, counts

def processar_zip(
    zip_file_path: Path,
    gerar_csv: bool = False,
) -> Tuple[str, List[Tuple[str, str, str]], Dict[str, int]]:
    """
    Compat: processa um único ZIP (modo antigo).
    """
    work_dir = Path(tempfile.mkdtemp())
    with zipfile.ZipFile(zip_file_path, "r") as z:
        z.extractall(work_dir)
    file_paths = [p for p in work_dir.rglob("*") if p.is_file()]
    return processar_arquivos(file_paths, gerar_csv=gerar_csv)

def processar_misto(
    paths: List[Path],
    gerar_csv: bool = False,
) -> Tuple[str, List[Tuple[str, str, str]], Dict[str, int]]:
    """
    NOVO: aceita arquivos “soltos” e/ou múltiplos .zip em uma mesma chamada.
    Expande cada .zip para uma pasta temporária e processa tudo de uma vez.
    """
    workspace = Path(tempfile.mkdtemp())
    coletados: List[Path] = []

    for p in paths:
        p = Path(p)
        if p.suffix.lower() == ".zip":
            sub = workspace / p.stem
            sub.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(p, "r") as z:
                z.extractall(sub)
            coletados.extend([q for q in sub.rglob("*") if q.is_file()])
        else:
            coletados.append(p)

    return processar_arquivos(coletados, gerar_csv=gerar_csv)

# ----------------------------------------------------------------------
# Resumo amigável
# ----------------------------------------------------------------------
def resumo_formatos(counts: Dict[str, int]) -> str:
    if not counts:
        return "Nenhuma extensão foi detectada."
    linhas = ["Foram encontrados:"]
    total = sum(counts.values())
    for ext, qtd in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
        desc, soft = FORMAT_INFO.get(ext, ("(sem descrição)", "—"))
        linhas.append(
            f"- {qtd} arquivo(s) com a extensão '{ext}' "
            f"({desc}; software recomendado: {soft})"
        )
    linhas.append(f"\nTotal de arquivos com extensão detectada: {total}")
    return "\n".join(linhas)
